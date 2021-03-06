# -*- coding: utf-8 -*-
import datetime
import json
import logging
from os import truncate
import openpyxl
import pathlib

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

def getExcelBookPath(prefix):
	def getExcelBookName():
		now = datetime.datetime.now()
		return prefix + now.strftime("_%Y%m%d_%H%M%S.xlsx")
	myPath = pathlib.Path(__file__)
	xmlRoot = myPath.parent
	return xmlRoot / 'resources' / getExcelBookName()

def getCrudConfig():
	myPath = pathlib.Path(__file__)
	json_path = myPath.parent / 'resources' / 'crudConfig.json'
	with json_path.open(mode='r', encoding='UTF8') as f:
		return json.load(f)

def formatCrudSheet(sheet, crud_config):
	excel_config = crud_config['Excel']
	start_row = excel_config['start_row']
	start_column = excel_config['start_column']
	crud_width = excel_config['crud_width']
	row = start_row
	col = start_column
	alignment1 = Alignment(horizontal="center", vertical="center")
	table_list = []
	for key, name in excel_config['tables'].items():
		logging.debug(key + ': ' + name)
		num = len(crud_config['tables'][key])
		d = sheet.cell(row=row, column=col, value=name)
		d.alignment = alignment1
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+num*4-1)
		count = 0
		for table in crud_config['tables'][key]:
			for k, v in table.items():
				table_list.append(k.lower())
				# テーブル名
				r = row + 1
				c = col + count * 4
				d = sheet.cell(row=r, column=c, value='%s\n%s' % (v,k))
				d.alignment = alignment1
				sheet.merge_cells(start_row=r, start_column=c, end_row=r, end_column=col+(count+1)*4-1)
				# insert
				r += 1
				a1 = openpyxl.utils.get_column_letter(c)
				sheet.column_dimensions[a1].width = crud_width
				d = sheet.cell(row=r, column=c,   value='C')
				d.alignment = alignment1
				# select
				c += 1
				a1 = openpyxl.utils.get_column_letter(c)
				sheet.column_dimensions[a1].width = crud_width
				d = sheet.cell(row=r, column=c, value='R')
				d.alignment = alignment1
				# update
				c += 1
				a1 = openpyxl.utils.get_column_letter(c)
				sheet.column_dimensions[a1].width = crud_width
				d = sheet.cell(row=r, column=c, value='U')
				d.alignment = alignment1
				# delete
				c += 1
				a1 = openpyxl.utils.get_column_letter(c)
				sheet.column_dimensions[a1].width = crud_width
				d = sheet.cell(row=r, column=c, value='D')
				d.alignment = alignment1
				count += 1

		col += num * 4
	return table_list

def _setCURD(sheet, row, start_column, alignment1, table_list, crud_info):
	for crud_type, crud_tables in crud_info.items():
		for table in crud_tables:
			table = table.lower()
			if table not in table_list:
				logging.debug('[%s] is not in CRUD list' % table)
				continue
			pos = start_column + table_list.index(table) * 4
			if crud_type == 'create':
				#pos += 0
				crud = 'C'
			elif crud_type == 'read':
				pos += 1
				crud = 'R'
			elif crud_type == 'update':
				pos += 2
				crud = 'U'
			elif crud_type == 'delete':
				pos += 3
				crud = 'D'
			else:
				assert False, "[%s] is not supported by Excel write process" % crud_type
			d = sheet.cell(row=row, column=pos, value=crud)
			d.alignment = alignment1

def formatClassMethods(sheet, crud_config, table_list, pinfo):
	excel_config = crud_config['Excel']
	start_row = excel_config['start_row']
	start_column = excel_config['start_column']
	row = start_row + excel_config['header_rows']
	exclude_no_mapper_call = excel_config['exclude-no_mapper_call']
	alignment1 = Alignment(horizontal="center", vertical="center")
	#クラス種別順で出力
	for class_type in ['controller', 'service', 'component', 'mapper' ,'other']:
		sheet.cell(row=row, column=1, value=class_type.capitalize())
		#クラスの出力順をFQCNの昇順で出力
		cl = sorted(pinfo[class_type].items(), key=lambda x:x[0])
		for c in cl:
			if exclude_no_mapper_call:
				# マッパー参照の有無をクラス単位に確認
				f = False
				for method_type in ['constructor', 'public-method', 'private-method']:
					if f:
						break
					for m in c[1]['methods'][method_type]:
						if m['mapper']:
							f = True
							break
				if not f:
					continue
			#クラス名
			sheet.cell(row=row, column=2, value=c[0])
			#メソッド種別順で出力
			for method_type in ['constructor', 'public-method', 'private-method']:
				#メソッド名の昇順で出力
				ml = sorted(c[1]['methods'][method_type], key=lambda x:x['name'])
				for m in ml:
					mpl = m['mapper']
					if exclude_no_mapper_call:
						# マッパー参照の有無をメソッド単位に確認
						if not mpl:
							continue
					#メソッド名
					sheet.cell(row=row, column=3, value=m['name'])
					#CRUD
					for mp in mpl:
						_setCURD(sheet, row, start_column, alignment1, table_list, mp['crud'])
					row += 1

def formatJavaSource(sheet, crud_config, table_list, pinfo):
	excel_config = crud_config['Excel']
	start_row = excel_config['start_row']
	start_column = excel_config['start_column']
	row = start_row + excel_config['header_rows']
	exclude_no_mapper_call = excel_config['exclude-no_mapper_call']
	alignment1 = Alignment(horizontal="center", vertical="center")
	prev_java_file = ''
	#クラス種別順で出力
	for class_type in ['controller', 'service', 'component', 'mapper' ,'other']:
		sheet.cell(row=row, column=1, value=class_type.capitalize())
		#クラスの出力順をFQCNの昇順で出力
		cl = sorted(pinfo[class_type].items(), key=lambda x:x[1]['source_path'])
		for c in cl:
			if exclude_no_mapper_call:
				# マッパー参照の有無をクラス単位に確認
				f = False
				for method_type in ['constructor', 'public-method', 'private-method']:
					if f:
						break
					for m in c[1]['methods'][method_type]:
						if m['mapper']:
							f = True
							break
				if not f:
					continue
			#ソースファイル名
			java_file = c[1]['source_path']
			sheet.cell(row=row, column=2, value=java_file)
			#メソッド種別順で出力
			for method_type in ['constructor', 'public-method', 'private-method']:
				#メソッド名の昇順で出力
				ml = sorted(c[1]['methods'][method_type], key=lambda x:x['name'])
				for m in ml:
					mpl = m['mapper']
					if exclude_no_mapper_call:
						# マッパー参照の有無をメソッド単位に確認
						if not mpl:
							continue
					#メソッド名
					#sheet.cell(row=row, column=3, value=m['name'])
					#CRUD
					for mp in mpl:
						_setCURD(sheet, row, start_column, alignment1, table_list, mp['crud'])

			if java_file != prev_java_file:
				row += 1
				prev_java_file = java_file

def setColumnWidth(sheet):
	def _setColumnWidth(col_no):
		max_width = 0
		columns = list(sheet.columns)[col_no-1]
		for cell in columns:
			if len(str(cell.value)) > max_width:
				max_width = len(str(cell.value))
		#adjusted_width = (max_width + 2) * 1.2
		adjusted_width = max_width
		sheet.column_dimensions[columns[0].column_letter].width = adjusted_width
	_setColumnWidth(1)    #クラスタイプ
	_setColumnWidth(2)	  #FQCN
	_setColumnWidth(3)	  #メソッド

def outExcel_method(crud_config, sheet, r):
	table_list = formatCrudSheet(sheet, crud_config)
	formatClassMethods(sheet, crud_config, table_list, r)

	setColumnWidth(sheet)

def outExcel_java(crud_config, sheet, r):
	table_list = formatCrudSheet(sheet, crud_config)
	formatJavaSource(sheet, crud_config, table_list, r)

	setColumnWidth(sheet)

def outExcel(r, view_info):
	crud_config = getCrudConfig()
	#print(crud_config)
	book = openpyxl.Workbook()
	# メソッド単位のCRUDを編集
	sheet = book.worksheets[0]
	sheet.title = 'crud methods'
	outExcel_method(crud_config, sheet, r)
	sheet =  book.create_sheet('crud java source', 0)
	outExcel_java(crud_config, sheet, r)

	book.save(getExcelBookPath(prefix='crud'))
	book.close()

def view2table(map_info, view_info):
	'''
	ビューの参照をテーブル参照に追加する
	'''
	for v in map_info.values():
		dml = v['dml']
		for info in dml.values():
			crudR = info['crud']['read']
			addTables = []
			for vn in crudR:
				if vn in view_info:
					for tn in view_info[vn]:
						if tn not in addTables:
							addTables.append(tn)
			for tn in addTables:
				if tn not in crudR:
					crudR.append(tn)

def outMapperInfo(map_info, view_info):
	crud_config = getCrudConfig()
	excel_config = crud_config['Excel']
	start_row = excel_config['start_row']
	start_column = excel_config['start_column']
	row = start_row + excel_config['header_rows']
	alignment1 = Alignment(horizontal="center", vertical="center")

	book = openpyxl.Workbook()
	sheet = book.worksheets[0]
	sheet.title = 'mapper'
	table_list = formatCrudSheet(sheet, crud_config)
	for dao in sorted(map_info.items(), key=lambda x:x[0]):
		sheet.cell(row=row, column=1, value=dao[0])
		for map_id in sorted(dao[1]['dml'].items(), key=lambda x:(x[1]['type'], x[0])):
			sheet.cell(row=row, column=2, value=map_id[1]['type'])
			sheet.cell(row=row, column=3, value=map_id[0])
			_setCURD(sheet, row, start_column, alignment1, table_list, map_id[1]['crud'])
			row += 1

	book.save(getExcelBookPath(prefix='mapper'))
	book.close()

if __name__ == '__main__':
	import pathlib

	from daoReader import DaoReader as DR
	from crudJudgment import judgment

	myPath = pathlib.Path(__file__)
	xmlRoot = myPath.parent

	reader = DR()
	r = reader.readXmls(xmlRoot = xmlRoot / 'resources')

	judgment(r)
	outExcel(r)
#[EOF]